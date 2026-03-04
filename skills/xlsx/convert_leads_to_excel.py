import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

csv_file = sys.argv[1]
xlsx_file = csv_file.replace('.csv', '_formatted.xlsx')

# Read CSV
df = pd.read_csv(csv_file)

# Reorder and select columns as requested
columns_order = [
    'Lead Title', 'Lead ID', 'Lead Created', 'Lead Updated', 
    'Expected Close Date', 'Status', 'Seen', 'Value', 'Currency',
    'Source', 'Origin',
    'Owner ID', 'Owner Name', 'Owner Email',
    'Creator ID', 'Creator Name',
    'Person ID', 'Person Name', 'Person Email', 'Person Phone',
    'Organization ID', 'Organization Name', 'Organization Address',
    'Next Activity ID', 'Visible To', 'CC Email', 'Label IDs'
]

# Filter to existing columns
existing_columns = [c for c in columns_order if c in df.columns]
df_export = df[existing_columns]

# Export to Excel
df_export.to_excel(xlsx_file, index=False, sheet_name='Leads')

# Apply formatting
wb = load_workbook(xlsx_file)
ws = wb.active

# Header styling
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Apply header format
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# Auto-fit columns
for column in ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    
    for cell in column:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    
    # Set width (min 12, max 50)
    adjusted_width = min(max_length + 2, 50)
    adjusted_width = max(adjusted_width, 12)
    ws.column_dimensions[column_letter].width = adjusted_width

# Freeze header row
ws.freeze_panes = 'A2'

# Apply borders to all cells
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = thin_border
        if cell.row > 1:
            cell.alignment = Alignment(vertical='top', wrap_text=False)

# Status color coding
status_col = existing_columns.index('Status') + 1 if 'Status' in existing_columns else None
if status_col:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=status_col)
        if cell.value == 'Active':
            cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        elif cell.value == 'Archived':
            cell.fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')

# Add summary sheet
summary_ws = wb.create_sheet('Summary')

summary_data = [
    ['Pipedrive Leads Export Summary'],
    [''],
    ['Total Leads', df_export.shape[0]],
    ['Columns', df_export.shape[1]],
    [''],
    ['Status Breakdown'],
    ['Status', 'Count'],
]

# Count by status
if 'Status' in df_export.columns:
    status_counts = df_export['Status'].value_counts()
    for status, count in status_counts.items():
        summary_data.append([status, count])

summary_data.extend([
    [''],
    ['Source Breakdown'],
    ['Source', 'Count'],
])

# Count by source
if 'Source' in df_export.columns:
    source_counts = df_export['Source'].value_counts()
    for source, count in source_counts.items():
        summary_data.append([source, count])

# Write summary
for row_idx, row_data in enumerate(summary_data, start=1):
    for col_idx, value in enumerate(row_data, start=1):
        cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
        
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif row_idx in [4, 7, 9, 12] and col_idx == 1:
            cell.font = Font(bold=True)

summary_ws.column_dimensions['A'].width = 25
summary_ws.column_dimensions['B'].width = 15

wb.save(xlsx_file)
print(f'Created: {xlsx_file}')
